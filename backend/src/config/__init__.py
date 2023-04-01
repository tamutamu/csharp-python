from enum import Enum
import openpyxl

from db.repository import SystemSettingRepository


class Const:
    class Status(str, Enum):
        PENDING = "PENDING"
        RUNNING = "RUNNING"
        WAITING = "WAITING"
        EXIT = "EXIT"

    class Result(str, Enum):
        SUCCESS = "SUCCESS"
        FAILED = "FAILED"
        SYSERR = "SYSERR"


class Config:
    DB_NAME = "main_db"
    AMAZON_PROFILE_NAME = "amazon"
    BACKEND_SERVER_PORT = -1
    FRONTEND_SERVER_PORT = -1
    SETTING_FILE_PATH = None

    MAX_RETRY = 3
    TIMEOUT_lv1 = 5
    TIMEOUT_lv2 = 3

    class Setting:
        AMAZON_USER_NAME = None
        AMAZON_USER_PASS = None
        PROFIT_RATE = 0
        FIXED_COST = 0
        USER_LIST = []


def load_setting():
    settings = SystemSettingRepository.dict()
    Config.SETTING_FILE_PATH = settings["SETTING_FILE_PATH"]

    user_list = []
    wb = openpyxl.load_workbook(Config.SETTING_FILE_PATH, data_only=True, read_only=True)
    ws = wb["ツール設定"]
    Config.Setting.AMAZON_USER_NAME = ws["B4"].value
    Config.Setting.AMAZON_USER_PASS = ws["B5"].value
    Config.Setting.PROFIT_RATE = int(ws["B6"].value)
    Config.Setting.FIXED_COST = int(ws["B7"].value)

    ws = wb["ユーザ情報"]
    for row in ws.iter_rows(min_row=4):
        if row[0].value in ("", None):
            break
        user_list.append([row[0].value, row[1].value, row[2].value])

    Config.Setting.USER_LIST = user_list


# mypy: ignore-errors
# class _const:
#     class ConstError(TypeError):
#         pass

#     def __setattr__(self, name, value):
#         if name in self.__dict__:
#             raise self.ConstError("Can't rebind const (%s)" % name)
#         self.__dict__[name] = value


# import sys

# sys.modules["config"] = _const()
